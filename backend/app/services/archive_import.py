from __future__ import annotations

import io
import json
import re
import urllib.error
import urllib.request
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path, PurePosixPath
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from pdfplumber import open as open_pdf
from rapidfuzz import fuzz, process
from PIL import Image
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models import ArchiveBatch, Partner, PriceDocument, PriceItem, Service
from app.services.pdf_document_ai import extract_rows_from_pdf_bytes as extract_pdf_document_ai_rows
from app.services.pdf_layout import LayoutWord, is_code_like, is_price_like, is_section_row, rows_from_words
from app.services.pdf_gemini_ocr import extract_rows_from_pdf_pages as extract_pdf_gemini_rows
from app.services.pdf_table_ocr import extract_pdf_table_rows_from_page, extract_pdf_table_rows_from_page_v2

try:
    import fitz  # type: ignore
except Exception:  # pragma: no cover
    fitz = None

from app.services.ocr import ocr_document_text


SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".xls"}
IGNORED_PATH_PARTS = {"__macosx", ".ds_store"}
DATE_PATTERNS = (
    re.compile(r"(?P<year>\d{4})[-_.](?P<month>\d{2})[-_.](?P<day>\d{2})"),
    re.compile(r"(?P<day>\d{2})[-_.](?P<month>\d{2})[-_.](?P<year>\d{4})"),
)
YEAR_PATTERN = re.compile(r"(?:^|[._\-\s])(?P<year>19\d{2}|20\d{2})(?:$|[._\-\s])")
NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
LEADING_CODE_RE = re.compile(
    r"^[\"'????\(\)\[\]\s]*([A-Z?-??0-9][A-Z?-??0-9.\-/]{1,24})\s+(.*)$",
    flags=re.IGNORECASE,
)

CODE_TRANSLATION = str.maketrans({
    "\u0410": "A", "\u0412": "B", "\u0415": "E", "\u041a": "K", "\u041c": "M", "\u041d": "H", "\u041e": "O", "\u041f": "P", "\u0420": "P", "\u0421": "C", "\u0422": "T", "\u0425": "X", "\u0423": "Y",
    "\u0430": "a", "\u0432": "b", "\u0435": "e", "\u043a": "k", "\u043c": "m", "\u043d": "h", "\u043e": "o", "\u043f": "p", "\u0440": "p", "\u0441": "c", "\u0442": "t", "\u0445": "x", "\u0443": "y",
})

HEADER_HINTS = ("name", "наим", "услуг", "service", "price", "cost", "цена", "стоим", "code", "код", "tariff")
REVIEW_PRICE_CHANGE_THRESHOLD = 0.5
FALLBACK_CURRENCY_RATES_TO_KZT = {
    "KZT": 1.0,
    "USD": 470.0,
    "EUR": 510.0,
    "RUB": 5.2,
}


@dataclass
class ArchiveImportStats:
    archive_name: str
    archive_id: str | None = None
    partner_count: int = 0
    document_count: int = 0
    item_count: int = 0
    matched_item_count: int = 0
    warnings: list[str] = field(default_factory=list)


def _clean_text(value: object | None) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\u00ad", "")
    return re.sub(r"\s+", " ", text).strip()


def _repair_mojibake(text: str) -> str:
    if not text:
        return text
    if not any(marker in text for marker in ("?", "?", "?", "?", "??", "?")):
        return text

    def score(candidate: str) -> tuple[int, int]:
        cyrillic = len(re.findall(r"[?-??-???]", candidate))
        mojibake = sum(candidate.count(marker) for marker in ("?", "?", "?", "?", "??", "?"))
        return cyrillic, -mojibake

    best = text
    best_score = score(text)
    for encoding in ("latin1", "cp1252", "cp1251"):
        try:
            candidate = text.encode(encoding).decode("utf-8")
        except Exception:
            continue
        candidate_score = score(candidate)
        if candidate_score > best_score:
            best = candidate
            best_score = candidate_score
    return best


def _truncate_text(value: object | None, limit: int) -> str | None:
    text = _clean_text(value)
    if not text:
        return None
    if len(text) <= limit:
        return text
    return text[: max(0, limit - 1)].rstrip() + "?"


def _row_needs_review(
    *,
    service_id: str | None,
    price_resident: float | None,
    price_nonresident: float | None,
    price_original: float | None,
) -> bool:
    if _price_validation_reasons(
        price_resident=price_resident,
        price_nonresident=price_nonresident,
        price_original=price_original,
    ):
        return True
    if service_id is None:
        return True
    return False


def _price_validation_reasons(
    *,
    price_resident: float | None,
    price_nonresident: float | None,
    price_original: float | None,
) -> list[str]:
    reasons: list[str] = []
    prices = [value for value in (price_resident, price_nonresident, price_original) if value is not None]
    if not prices:
        reasons.append("price missing or invalid")
    if any(value <= 0 for value in prices):
        reasons.append("price must be > 0")
    if (
        price_resident is not None
        and price_nonresident is not None
        and price_nonresident < price_resident
    ):
        reasons.append("nonresident price is lower than resident price")
    return reasons


def _detect_currency(text: str | None) -> str:
    normalized = _clean_text(text).casefold()
    if any(token in normalized for token in ("usd", "$", "доллар")):
        return "USD"
    if any(token in normalized for token in ("eur", "€", "евро")):
        return "EUR"
    if any(token in normalized for token in ("rub", "руб")):
        return "RUB"
    return "KZT"


@lru_cache(maxsize=32)
def _exchange_rate_to_kzt(currency: str) -> tuple[float, str]:
    currency = currency.upper()
    if currency == "KZT":
        return 1.0, "native"
    url = settings.exchange_rate_api_url.format(currency=currency)
    try:
        with urllib.request.urlopen(url, timeout=settings.exchange_rate_timeout_seconds) as response:
            payload = json.loads(response.read().decode("utf-8"))
        rate = payload.get("rates", {}).get("KZT")
        if isinstance(rate, (int, float)) and rate > 0:
            return float(rate), "api"
    except (urllib.error.URLError, TimeoutError, ValueError, json.JSONDecodeError, KeyError):
        pass
    return FALLBACK_CURRENCY_RATES_TO_KZT.get(currency, 1.0), "fallback"


def _convert_price_to_kzt(
    value: float | None,
    currency: str,
    stats: ArchiveImportStats,
) -> tuple[float | None, str]:
    if value is None:
        return None, "native"
    rate, source = _exchange_rate_to_kzt(currency)
    if source == "fallback" and currency != "KZT":
        warning = f"Exchange rate API unavailable for {currency}; fallback rate used"
        if warning not in stats.warnings:
            stats.warnings.append(warning)
    return round(float(value) * rate, 2), source


def _clean_partner_name_from_stem(stem: str) -> str:
    cleaned = re.sub(r"(прайс|price|год)", " ", stem, flags=re.IGNORECASE)
    cleaned = YEAR_PATTERN.sub(" ", cleaned)
    cleaned = re.sub(r"[_-]+", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ._-")
    return cleaned or stem


def _extract_date(text: str) -> date | None:
    for pattern in DATE_PATTERNS:
        match = pattern.search(text)
        if not match:
            continue
        try:
            return date(
                int(match.group("year")),
                int(match.group("month")),
                int(match.group("day")),
            )
        except ValueError:
            continue
    year_match = YEAR_PATTERN.search(text)
    if year_match:
        try:
            return date(int(year_match.group("year")), 1, 1)
        except ValueError:
            return None
    return None


def _normalize_path(path: str) -> PurePosixPath:
    normalized = PurePosixPath(path.replace("\\", "/"))
    parts = [part for part in normalized.parts if part not in (".", "")]
    return PurePosixPath(*parts)


def _group_key(entry_path: PurePosixPath) -> str:
    if len(entry_path.parts) > 1:
        return entry_path.parts[0]
    return "__root__"


def _choose_partner_name(archive_name: str, group_key: str) -> str:
    if group_key != "__root__":
        return _clean_partner_name_from_stem(group_key.replace("_", " ").strip()) or archive_name
    return archive_name


def _extract_pdf_text_and_tables(data: bytes) -> tuple[list[list[str]], str]:
    rows: list[list[str]] = []
    text_chunks: list[str] = []
    price_pattern = re.compile(r"(?P<price>\d[\d\s]*\d|\d)\s*$")
    with open_pdf(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            page_text = _clean_text(page.extract_text())
            if page_text:
                text_chunks.append(page_text)
            words = page.extract_words(keep_blank_chars=False, use_text_flow=True)
            if not words:
                continue

            bands: list[tuple[float, list[dict[str, object]]]] = []
            for word in sorted(words, key=lambda w: (float(w["top"]), float(w["x0"]))):
                top = float(word["top"])
                if not bands or abs(top - bands[-1][0]) > 4:
                    bands.append((top, [word]))
                else:
                    bands[-1][1].append(word)

            pending = ""
            for _top, band_words in bands:
                line = " ".join(str(w["text"]) for w in sorted(band_words, key=lambda w: float(w["x0"])))
                line = _clean_text(line)
                if not line:
                    continue
                lowered = line.lower()
                if lowered.startswith("раздел") or ("код" in lowered and "наименование" in lowered) or ("стоим" in lowered and "тенге" in lowered):
                    pending = ""
                    continue

                combined = f"{pending} {line}".strip() if pending else line
                if price_pattern.search(combined):
                    price_match = price_pattern.search(combined)
                    if not price_match:
                        continue
                    price = price_match.group("price")
                    body = combined[: price_match.start()].strip(" ,.;:-")
                    pending = ""
                    if not body:
                        continue

                    tokens = body.split()
                    code = ""
                    name = body
                    if tokens:
                        if len(tokens) >= 2 and len(tokens[0]) <= 2 and re.match(r"^\d[\d.]*$", tokens[1]):
                            code = f"U{tokens[1]}"
                            name = " ".join(tokens[2:]).strip(" ,.;:-")
                        else:
                            first = tokens[0]
                            if re.match(r"^[A-ZА-ЯЁІЇ0-9][A-ZА-ЯЁІЇ0-9.\-]*$", first, flags=re.IGNORECASE):
                                code = first.replace("и", "U").replace("І", "I").replace("Ї", "I")
                                name = " ".join(tokens[1:]).strip(" ,.;:-")
                    if not name:
                        continue
                    rows.append([code, name, price])
                else:
                    pending = combined
    return rows, "\n".join(text_chunks)


def _extract_pdf_table_rows(data: bytes, stats: ArchiveImportStats) -> tuple[list[list[str]], str]:
    if fitz is None:
        stats.warnings.append("PDF table OCR skipped: PyMuPDF is not available")
        return [], ""

    rows: list[list[str]] = []
    text_chunks: list[str] = []
    with fitz.open(stream=data, filetype="pdf") as pdf:  # type: ignore[union-attr]
        for page in pdf:
            page_rows, page_text = extract_pdf_table_rows_from_page(page, stats)
            rows.extend(page_rows)
            if page_text:
                text_chunks.append(page_text)
    return rows, "\n".join(text_chunks)


def _extract_pdf_text_and_tables_v2(data: bytes) -> tuple[list[list[str]], str]:
    rows: list[list[str]] = []
    with open_pdf(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            page_rows, _ = _extract_pdf_text_rows_from_page(page)
            rows.extend(page_rows)

    text_chunks = [" | ".join(cell for cell in row if cell) for row in rows if any(cell for cell in row)]
    return rows, "\n".join(text_chunks)


def _extract_pdf_text_rows_from_page(page: object) -> tuple[list[list[str]], str]:
    raw_words = page.get_text("words")
    words = [
        LayoutWord(text=str(word[4] or ""), x0=float(word[0]), y0=float(word[1]), x1=float(word[2]), y1=float(word[3]))
        for word in raw_words
        if len(word) >= 5 and str(word[4] or "").strip()
    ]
    if not words:
        return [], ""
    rows = rows_from_words(words, page_width=getattr(page, "width", None))
    text_chunks = [" | ".join(cell for cell in row if cell) for row in rows if any(cell for cell in row)]
    return rows, "\n".join(text_chunks)


def _extract_pdf_ocr_text_rows_from_page(page: object, stats: ArchiveImportStats) -> tuple[list[list[str]], str]:
    if fitz is None:
        return [], ""
    pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)  # type: ignore[union-attr]
    image_bytes = pixmap.tobytes("png")
    image = Image.open(io.BytesIO(image_bytes))
    try:
        text, engine = ocr_document_text(image, engine="auto")
        if engine == "vision":
            stats.warnings.append("OCR used Google Cloud Vision")
    except Exception as exc:  # pragma: no cover
        stats.warnings.append(f"OCR failed for one PDF page: {exc}")
        return [], ""

    rows: list[list[str]] = []
    text_chunks: list[str] = []
    for line in text.splitlines():
        cleaned = _clean_text(line)
        if not cleaned:
            continue
        rows.append([cleaned])
        text_chunks.append(cleaned)
    return rows, "\n".join(text_chunks)


def _score_pdf_rows(rows: list[list[str]]) -> int:
    score = 0
    for row in rows:
        cleaned = [_clean_text(cell) for cell in row if _clean_text(cell)]
        if not cleaned:
            continue
        score += len(cleaned)
        if len(cleaned) >= 3:
            score += 3
        if len(cleaned) == 1:
            score -= 1
        joined = " ".join(cleaned)
        if len(joined) > 160:
            score -= 2
        if re.search(r"\d[\d\s]*\d\s*$", joined):
            score += 2
        if re.search(r"^[A-ZА-Я0-9][A-ZА-Я0-9.\-]*\s+", joined):
            score += 1
    return score


def _score_pdf_rows_v2(rows: list[list[str]]) -> int:
    score = 0
    for row in rows:
        cleaned = [_clean_text(cell) for cell in row if _clean_text(cell)]
        if not cleaned:
            continue
        score += len(cleaned)
        if 2 <= len(cleaned) <= 6:
            score += 3
        if len(cleaned) == 1:
            if is_section_row(cleaned[0]):
                score += 1
            else:
                score -= 2
        joined = " ".join(cleaned)
        if len(joined) > 180:
            score -= 2
        if any(is_code_like(cell) for cell in cleaned):
            score += 2
        if any(is_price_like(cell) for cell in cleaned):
            score += 1
        if is_section_row(joined):
            score += 1
    return score


def _render_pdf_page_image(page: object, dpi: int | None = None) -> Image.Image:
    if fitz is None:
        raise RuntimeError("PyMuPDF is not available")
    resolved_dpi = dpi or settings.pdf_ocr_dpi
    scale = resolved_dpi / 72.0
    pixmap = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)  # type: ignore[union-attr]
    return Image.frombytes("RGB", [pixmap.width, pixmap.height], pixmap.samples)


def _ocr_pdf_text(data: bytes, stats: ArchiveImportStats) -> tuple[list[list[str]], str]:
    if fitz is None:
        stats.warnings.append("OCR skipped: PyMuPDF is not available")
        return [], ""

    rows: list[list[str]] = []
    text_chunks: list[str] = []
    with fitz.open(stream=data, filetype="pdf") as pdf:  # type: ignore[union-attr]
        for page in pdf:
            pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)  # type: ignore[union-attr]
            image_bytes = pixmap.tobytes("png")
            image = Image.open(io.BytesIO(image_bytes))
            try:
                text, engine = ocr_document_text(image, engine="auto")
                text = _clean_text(text)
                if engine == "vision":
                    stats.warnings.append("OCR used Google Cloud Vision")
            except Exception as exc:  # pragma: no cover
                stats.warnings.append(f"OCR failed for one PDF page: {exc}")
                text = ""
            if text:
                text_chunks.append(text)
                for line in text.splitlines():
                    cleaned = _clean_text(line)
                    if cleaned:
                        rows.append([cleaned])
    return rows, "\n".join(text_chunks)


def _extract_docx_rows(data: bytes) -> tuple[list[list[str]], str]:
    rows: list[list[str]] = []
    text_chunks: list[str] = []
    with zipfile.ZipFile(io.BytesIO(data)) as docx_zip:
        xml = docx_zip.read("word/document.xml")
    root = ET.fromstring(xml)

    def visible_text(node: ET.Element) -> str:
        local_name = node.tag.rsplit("}", 1)[-1]
        if local_name == "del":
            return ""
        if local_name == "t":
            return node.text or ""
        if local_name in {"tab", "br", "cr"}:
            return "\n"
        parts: list[str] = []
        if node.text:
            parts.append(node.text)
        for child in list(node):
            parts.append(visible_text(child))
            if child.tail:
                parts.append(child.tail)
        return "".join(parts)

    for paragraph in root.findall(".//w:p", NS):
        text = _clean_text(visible_text(paragraph))
        if text:
            text_chunks.append(text)

    for table in root.findall(".//w:tbl", NS):
        for row in table.findall("./w:tr", NS):
            cleaned: list[str] = []
            for cell in row.findall("./w:tc", NS):
                cell_text = _clean_text(visible_text(cell))
                cleaned.append(cell_text)
            if any(cleaned):
                rows.append(cleaned)

    return rows, "\n".join(text_chunks)


def _extract_xlsx_rows(data: bytes) -> tuple[list[list[str]], str]:
    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    rows: list[list[str]] = []
    sheet_names: list[str] = []
    for sheet in workbook.worksheets:
        sheet_names.append(sheet.title)
        for raw_row in sheet.iter_rows(values_only=True):
            cleaned = [_clean_text(cell) for cell in raw_row]
            if any(cleaned):
                rows.append(cleaned)
    return rows, ", ".join(sheet_names)


def _extract_xls_rows(data: bytes) -> tuple[list[list[str]], str]:
    try:
        import xlrd  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("Reading XLS requires xlrd") from exc

    workbook = xlrd.open_workbook(file_contents=data)
    rows: list[list[str]] = []
    sheet_names: list[str] = []
    for sheet in workbook.sheets():
        sheet_names.append(sheet.name)
        for row_idx in range(sheet.nrows):
            cleaned: list[str] = []
            for cell in sheet.row(row_idx):
                if cell.ctype == xlrd.XL_CELL_ERROR:
                    cleaned.append(xlrd.error_text_from_code.get(cell.value, "#ERROR"))
                    continue
                cleaned.append(_clean_text(cell.value))
            if any(cleaned):
                rows.append(cleaned)
    return rows, ", ".join(sheet_names)


def extract_rows_and_content(file_path: Path, data: bytes, stats: ArchiveImportStats) -> tuple[list[list[str]], str]:
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        if fitz is None:
            return _ocr_pdf_text(data, stats)

        if settings.ocr_provider.strip().lower() == "document_ai" or settings.document_ai_processor_name.strip() or settings.document_ai_auto_create_processor:
            try:
                docai_rows, docai_text = extract_pdf_document_ai_rows(data, stats)
                if docai_rows or docai_text:
                    return docai_rows, docai_text
            except Exception as exc:
                stats.warnings.append(f"Document AI PDF OCR failed: {exc}")

        gemini_enabled = bool(settings.gemini_api_key.strip())
        page_rows: dict[int, list[list[str]]] = {}
        page_text_chunks: dict[int, str] = {}
        local_rows_by_page: dict[int, list[list[str]]] = {}
        local_text_by_page: dict[int, str] = {}
        gemini_pages: list[tuple[int, Image.Image]] = []
        fallback_pages: list[tuple[int, Image.Image]] = []
        with fitz.open(stream=data, filetype="pdf") as pdf:  # type: ignore[union-attr]
            for page_index, page in enumerate(pdf, start=1):
                if gemini_enabled:
                    try:
                        gemini_pages.append((page_index, _render_pdf_page_image(page)))
                        continue
                    except Exception as exc:
                        stats.warnings.append(f"Failed to render PDF page {page_index} for Gemini OCR: {exc}")
                        try:
                            fallback_pages.append((page_index, _render_pdf_page_image(page, dpi=180)))
                        except Exception as render_exc:
                            stats.warnings.append(f"Failed to render PDF page {page_index} for fallback OCR: {render_exc}")
                            continue
                    continue

                text_rows, text_content = _extract_pdf_text_rows_from_page(page)
                text_score = _score_pdf_rows_v2(text_rows)
                table_rows, table_content = extract_pdf_table_rows_from_page_v2(page, stats)
                table_score = _score_pdf_rows_v2(table_rows)
                use_table = bool(table_rows) and table_score > text_score and (
                    table_score - text_score >= 20 or text_score < 100
                )
                local_rows = table_rows if use_table else text_rows
                local_text = table_content if use_table else text_content
                if local_rows:
                    local_rows_by_page[page_index] = local_rows
                if local_text:
                    local_text_by_page[page_index] = local_text

                if local_rows and _score_pdf_rows_v2(local_rows) >= 120:
                    page_rows[page_index] = local_rows
                    if local_text:
                        page_text_chunks[page_index] = local_text
                    continue

                try:
                    fallback_pages.append((page_index, _render_pdf_page_image(page, dpi=180)))
                except Exception as render_exc:
                    stats.warnings.append(f"Failed to render PDF page {page_index} for fallback OCR: {render_exc}")
                    continue

        if gemini_pages:
            try:
                gemini_rows_by_page = extract_pdf_gemini_rows(gemini_pages, stats)
                for page_index, rows_for_page in gemini_rows_by_page.items():
                    if rows_for_page:
                        page_rows[page_index] = rows_for_page
                        page_text_chunks[page_index] = "\n".join(
                            " | ".join(cell for cell in row if cell) for row in rows_for_page if any(cell for cell in row)
                        )
            except Exception as exc:
                stats.warnings.append(f"Gemini PDF OCR failed: {exc}")

        for page_index, rows_for_page in local_rows_by_page.items():
            if page_index not in page_rows and rows_for_page:
                page_rows[page_index] = rows_for_page
                text_chunk = local_text_by_page.get(page_index)
                if text_chunk:
                    page_text_chunks[page_index] = text_chunk

        if fallback_pages:
            for page_index, image in fallback_pages:
                try:
                    text, engine = ocr_document_text(image, engine="auto")
                    text = _clean_text(text)
                    if engine == "vision":
                        stats.warnings.append("OCR used Google Cloud Vision")
                except Exception as exc:  # pragma: no cover
                    stats.warnings.append(f"OCR failed for PDF page {page_index}: {exc}")
                    continue
                if text:
                    page_rows.setdefault(page_index, [])
                    page_rows[page_index].extend([[line] for line in text.splitlines() if _clean_text(line)])
                    page_text_chunks[page_index] = text

        if not page_rows and not page_text_chunks:
            return _ocr_pdf_text(data, stats)
        flattened_rows: list[list[str]] = []
        flattened_text: list[str] = []
        for page_index in sorted(set(page_rows) | set(page_text_chunks)):
            rows_for_page = page_rows.get(page_index, [])
            if rows_for_page:
                flattened_rows.extend(rows_for_page)
            text_chunk = page_text_chunks.get(page_index)
            if text_chunk:
                flattened_text.append(text_chunk)
        return flattened_rows, "\n".join(flattened_text)
    if suffix == ".docx":
        return _extract_docx_rows(data)
    if suffix == ".xlsx":
        return _extract_xlsx_rows(data)
    if suffix == ".xls":
        return _extract_xls_rows(data)
    raise ValueError(f"Unsupported file type: {suffix}")


def _build_service_choices(session: Session) -> list[tuple[str, str, str]]:
    choices: list[tuple[str, str, str]] = []
    services = session.scalars(select(Service).where(Service.is_active.is_(True))).all()
    for service in services:
        choices.append((service.service_name, str(service.service_id), "service"))
        for synonym in service.synonyms or []:
            synonym_text = _clean_text(synonym)
            if synonym_text:
                choices.append((synonym_text, str(service.service_id), "synonym"))
    return choices


def _best_service_match(raw_name: str, choices: list[tuple[str, str, str]]) -> tuple[str | None, int, str]:
    if not raw_name or not choices:
        return None, 0, "none"
    if _looks_uncertain(raw_name):
        return None, 0, "none"
    normalized_raw = _clean_text(raw_name).casefold()
    for candidate_label, service_id, match_source in choices:
        if _clean_text(candidate_label).casefold() == normalized_raw:
            if match_source == "synonym":
                return service_id, 100, "synonym"
            return service_id, 100, "exact"
    labels = [label for label, _service_id, _match_source in choices]
    match = process.extractOne(raw_name, labels, scorer=fuzz.token_sort_ratio)
    if not match:
        return None, 0, "none"
    label, score, _ = match
    if score >= settings.service_match_auto_threshold:
        for candidate_label, service_id, _match_source in choices:
            if candidate_label == label:
                return service_id, score, "fuzzy_auto"
        return None, score, "none"
    if score >= settings.service_match_review_threshold:
        return None, score, "fuzzy_review"
    for candidate_label, service_id, _match_source in choices:
        if candidate_label == label:
            return None, score, "none"
    return None, score, "none"


def _looks_uncertain(raw_name: str) -> bool:
    text = _clean_text(raw_name).lower()
    if not text:
        return True
    if len(text) < 4:
        return True
    if any(token in text for token in ("стоимость", "наименование услуги", "раздел", "приложение", "таблица")):
        return True
    letters = sum(1 for char in text if char.isalpha())
    digits = sum(1 for char in text if char.isdigit())
    if letters < 3:
        return True
    if digits and digits / max(len(text), 1) > 0.35:
        return True
    return False


def _resolve_service_name(row: list[str], header: list[str]) -> str | None:
    normalized_header = [cell.lower() for cell in header]
    preferred_columns = ("name", "наим", "услуг", "service", "description", "опис")
    for index, cell_name in enumerate(normalized_header):
        if any(token in cell_name for token in preferred_columns):
            candidate = _clean_text(row[index] if index < len(row) else None)
            if candidate:
                return candidate
    for value in row:
        candidate = _clean_text(value)
        if candidate:
            return candidate
    return None


def _resolve_code(row: list[str], header: list[str]) -> str | None:
    normalized_header = [cell.lower() for cell in header]
    for index, cell_name in enumerate(normalized_header):
        if any(token in cell_name for token in ("code", "код", "tariff", "id")):
            candidate = _clean_text(row[index] if index < len(row) else None)
            if candidate:
                return candidate
    return None


def _resolve_price(row: list[str], header: list[str]) -> float | None:
    normalized_header = [cell.lower() for cell in header]
    for index, cell_name in enumerate(normalized_header):
        if any(token in cell_name for token in ("price", "cost", "цена", "стоим", "tariff", "sum")):
            candidate = _clean_text(row[index] if index < len(row) else None)
            price = _parse_price_value(candidate)
            if price is not None:
                return price
    for value in row:
        price = _parse_price_value(value)
        if price is not None:
            return price
    return None


def _parse_price_value(value: str | None) -> float | None:
    if not value:
        return None
    normalized = value.replace("\xa0", " ").replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", normalized)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _split_leading_code_and_name(value: str | None) -> tuple[str | None, str | None]:
    text = _clean_text(value)
    if not text:
        return None, None
    text = text.strip(" ,.;:-|")
    if not text or _parse_price_value_v2(text) is not None:
        return None, text

    tokens = text.split()
    for index, token in enumerate(tokens):
        normalized = token.translate(CODE_TRANSLATION)
        if not is_code_like(normalized):
            continue
        remainder = _clean_text(" ".join(tokens[index + 1 :]))
        if not remainder or is_price_like(remainder):
            continue
        return normalized.replace(" ", ""), remainder

    return None, text


def _header_score(row: list[str]) -> int:
    score = 0
    for cell in row:
        lowered = cell.lower()
        if any(token in lowered for token in HEADER_HINTS):
            score += 1
    return score


def _pick_header_index(rows: list[list[str]]) -> int | None:
    best_index: int | None = None
    best_score = 0
    for index, row in enumerate(rows[:10]):
        if not any(_clean_text(cell) for cell in row):
            continue
        score = _header_score(row)
        if score > best_score:
            best_index = index
            best_score = score
    return best_index if best_score > 0 else None


def _parse_price_value_v2(value: str | None) -> float | None:
    if not value:
        return None
    normalized = _clean_text(value).lower()
    normalized = normalized.replace(" ", " ")
    normalized = normalized.replace("?????", "").replace("kzt", "").replace("??", "").replace("?", "")
    if re.search(r"[a-z?-?]", normalized, flags=re.IGNORECASE):
        return None
    normalized = normalized.replace(" ", "").replace(",", ".")
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", normalized):
        return None
    try:
        parsed = float(normalized)
    except ValueError:
        return None
    if abs(parsed) < 10:
        return None
    if abs(parsed) >= 10_000_000_000:
        return None
    return parsed


def _header_score(row: list[str]) -> int:
    score = 0
    for cell in row:
        lowered = cell.lower()
        if any(token in lowered for token in HEADER_HINTS):
            score += 1
    return score


def _pick_header_index(rows: list[list[str]]) -> int | None:
    best_index: int | None = None
    best_score = 0
    for index, row in enumerate(rows[:10]):
        if not any(_clean_text(cell) for cell in row):
            continue
        score = _header_score(row)
        if score > best_score:
            best_index = index
            best_score = score
    return best_index if best_score > 0 else None


def _parse_price_value_v2(value: str | None) -> float | None:
    if not value:
        return None
    normalized = _clean_text(value).lower()
    normalized = normalized.replace("\xa0", " ").replace("тенге", "").replace("kzt", "").replace("тг", "")
    numeric_match = re.search(r"\d[\d\s]*(?:[.,]\d+)?", normalized)
    if not numeric_match:
        return None
    normalized = numeric_match.group(0).replace(" ", "").replace(",", ".")
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", normalized):
        return None
    try:
        parsed = float(normalized)
    except ValueError:
        return None
    if abs(parsed) < 10:
        return None
    if abs(parsed) >= 10_000_000_000:
        return None
    return parsed


def _header_score(row: list[str]) -> int:
    score = 0
    for cell in row:
        lowered = _clean_text(cell).lower()
        if not lowered:
            continue
        if any(token in lowered for token in HEADER_HINTS):
            score += 2
        if any(token in lowered for token in ("наимен", "услуг", "ед.", "ед ", "измер", "кол-", "ндс")):
            score += 1
    return score


def _pick_header_index(rows: list[list[str]]) -> int | None:
    best_index: int | None = None
    best_score = 0
    for index, row in enumerate(rows[:50]):
        if not any(_clean_text(cell) for cell in row):
            continue
        score = _header_score(row)
        if score > best_score:
            best_index = index
            best_score = score
    return best_index if best_score >= 2 else None


def _header_name_indices(header: list[str]) -> list[int]:
    indices: list[int] = []
    for index, cell_name in enumerate(header):
        lowered = _clean_text(cell_name).lower()
        if any(token in lowered for token in ("наимен", "услуг", "service", "name", "опис")):
            indices.append(index)
    return indices


def _row_price_values(row: list[str], header: list[str]) -> list[float]:
    normalized_header = [cell.lower() for cell in header]
    price_indices = [
        index
        for index, cell_name in enumerate(normalized_header)
        if any(token in cell_name for token in ("price", "cost", "С†РµРЅР°", "СЃС‚РѕРёРј", "tariff", "sum"))
    ]
    values: list[float] = []
    if price_indices:
        for index in price_indices:
            candidate = _clean_text(row[index] if index < len(row) else None)
            price = _parse_price_value_v2(candidate)
            if price is not None:
                values.append(price)
        if values:
            return values

    numeric_cells = [value for value in row if re.search(r"\d", _clean_text(value))]
    if len(numeric_cells) < 2:
        return []

    for value in numeric_cells:
        price = _parse_price_value_v2(value)
        if price is not None:
            values.append(price)
    return values


def _resolve_service_name_v2(row: list[str], header: list[str]) -> str | None:
    normalized_header = [cell.lower() for cell in header]
    preferred_columns = ("name", "service", "description", "?????", "??????")
    for index, cell_name in enumerate(normalized_header):
        if any(token in cell_name for token in preferred_columns):
            candidate = _clean_text(row[index] if index < len(row) else None)
            if candidate:
                text = candidate.strip(" ,.;:-|")
                if text and _parse_price_value_v2(text) is None:
                    tokens = text.split()
                    for token_index, token in enumerate(tokens):
                        normalized = token.translate(CODE_TRANSLATION)
                        if not is_code_like(normalized):
                            continue
                        remainder = _clean_text(" ".join(tokens[token_index + 1 :]))
                        if remainder and not is_price_like(remainder):
                            return remainder
                return candidate
    candidates: list[str] = []
    for value in row:
        candidate = _clean_text(value)
        if not candidate or is_code_like(candidate) or is_price_like(candidate):
            continue
        text = candidate.strip(" ,.;:-|")
        if text and _parse_price_value_v2(text) is None:
            tokens = text.split()
            for token_index, token in enumerate(tokens):
                normalized = token.translate(CODE_TRANSLATION)
                if not is_code_like(normalized):
                    continue
                remainder = _clean_text(" ".join(tokens[token_index + 1 :]))
                if remainder and not is_price_like(remainder):
                    return remainder
        candidates.append(candidate)
    if candidates:
        candidates.sort(key=len, reverse=True)
        return candidates[0]
    return None


def _resolve_code_v2(row: list[str], header: list[str]) -> str | None:
    normalized_header = [cell.lower() for cell in header]
    for index, cell_name in enumerate(normalized_header):
        if any(token in cell_name for token in ("code", "tariff", "id", "???", "???????")):
            candidate = _clean_text(row[index] if index < len(row) else None)
            if candidate:
                normalized = candidate.replace(" ", "").translate(CODE_TRANSLATION)
                if re.match(r"^[UYI]?\d", normalized):
                    normalized = normalized.replace("?", "U").replace("?", "U")
                return normalized
    for value in row:
        candidate = _clean_text(value)
        if not candidate or _parse_price_value_v2(candidate) is not None:
            continue
        text = candidate.strip(" ,.;:-|")
        if text and _parse_price_value_v2(text) is None:
            tokens = text.split()
            for token_index, token in enumerate(tokens):
                normalized = token.translate(CODE_TRANSLATION)
                if not is_code_like(normalized):
                    continue
                remainder = _clean_text(" ".join(tokens[token_index + 1 :]))
                if remainder and not is_price_like(remainder):
                    return normalized.replace(" ", "")
        normalized = candidate.replace(" ", "").translate(CODE_TRANSLATION)
        if re.match(r"^[UYI]?\d", normalized):
            normalized = normalized.replace("?", "U").replace("?", "U")
        if is_code_like(normalized):
            return normalized
    return None


def _resolve_prices_v2(row: list[str], header: list[str]) -> tuple[float | None, float | None, float | None]:
    values = _row_price_values(row, header)
    if not values:
        return None, None, None
    if len(values) == 1:
        return values[0], None, values[0]
    if len(values) == 2:
        return values[0], values[1], values[0]
    return values[0], values[1], values[0]

def _normalize_code_token(token: str) -> str:
    normalized = token.translate(CODE_TRANSLATION).replace(" ", "")
    normalized = normalized.replace("?", "U").replace("?", "U").replace("?", "U").replace("?", "U")
    normalized = normalized.replace("?", "U").replace("?", "U")
    return normalized


def _strip_leading_row_noise(text: str) -> str:
    cleaned = _clean_text(text)
    cleaned = re.sub(r"^[ий]\s+\d+(?:\.\d+)+\s+", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"^\d+(?:\.\d+)+\s+", "", cleaned)
    return cleaned.strip(" ,.;:-|")


def _extract_code_from_row(row: list[str]) -> str | None:
    has_alpha = any(any(ch.isalpha() for ch in _clean_text(cell)) for cell in row)
    for value in row:
        candidate = _clean_text(value)
        if not candidate or is_section_row(candidate):
            continue
        tokens = candidate.strip(" ,.;:-|").split()
        for token in tokens:
            normalized = _normalize_code_token(token)
            if not normalized:
                continue
            if re.fullmatch(r"\d+(?:\.\d+)+", normalized) and has_alpha:
                return f"U{normalized}"
            if is_price_like(normalized):
                continue
            if is_code_like(normalized):
                return normalized
    return None


def _extract_price_values_from_row(row: list[str], header: list[str]) -> list[float]:
    values = _row_price_values(row, header)
    if values:
        return values

    fallback: list[float] = []
    start_index = 0
    if header:
        name_indices = _header_name_indices(header)
        if name_indices:
            start_index = max(name_indices) + 1
        elif len(row) >= 4:
            start_index = 4
    elif len(row) >= 4:
        start_index = 3

    for value in row[start_index:]:
        price = _parse_price_value_v2(value)
        if price is not None:
            fallback.append(price)
    return fallback


def _extract_service_name_from_row(row: list[str], code: str | None, prices: list[float], header: list[str]) -> str | None:
    cleaned = [_clean_text(cell) for cell in row if _clean_text(cell)]
    if not cleaned:
        return None
    row_text = " ".join(cleaned)
    if is_section_row(row_text):
        return None

    pieces: list[str] = []
    for cell in cleaned:
        normalized_cell = _clean_text(cell)
        if code and _normalize_code_token(normalized_cell) == code:
            continue
        if is_price_like(normalized_cell):
            continue
        text = _strip_leading_row_noise(normalized_cell)
        if code and text.startswith(code):
            text = _clean_text(text[len(code):])
        else:
            tokens = text.split()
            if tokens:
                first = _normalize_code_token(tokens[0])
                if first == code or is_code_like(first):
                    text = _clean_text(" ".join(tokens[1:]))
        if text and not is_price_like(text):
            pieces.append(text)

    name = _clean_text(" ".join(pieces))
    if not name:
        return None
    if code and name == code:
        return None
    if len(name) < 3:
        return None
    if re.fullmatch(r"[\d\s.,-]+", name):
        return None
    return name


def _rows_from_source(rows: list[list[str]]) -> list[dict[str, str | float | None]]:
    if not rows:
        return []

    header_index = _pick_header_index(rows)
    header = rows[header_index] if header_index is not None else []
    data_rows = rows[header_index + 1 :] if header_index is not None else rows
    parsed: list[dict[str, str | float | None]] = []

    for row in data_rows:
        cleaned = [_clean_text(cell) for cell in row if _clean_text(cell)]
        if not cleaned:
            continue
        row_text = " ".join(cleaned)
        if is_section_row(row_text):
            continue

        code = _extract_code_from_row(cleaned)
        prices = _extract_price_values_from_row(cleaned, header)
        service_name = _extract_service_name_from_row(cleaned, code, prices, header)

        if not service_name and not code and not prices:
            continue
        if service_name and re.match(r"^\d", service_name):
            if not prices:
                continue
            service_name = _strip_leading_row_noise(service_name)
        if not service_name:
            # Keep only clearly structured rows; avoid junk rows with just a code or price.
            if not prices or len(prices) == 0:
                continue
            service_name = code or row_text

        resident_price = prices[0] if prices else None
        nonresident_price = prices[1] if len(prices) > 1 else None
        original_price = resident_price
        parsed.append(
            {
                "service_name_raw": service_name,
                "service_code_source": code,
                "price_resident_kzt": resident_price,
                "price_nonresident_kzt": nonresident_price,
                "price_original": original_price,
                "raw_row": row_text,
            }
        )
    return parsed


def _row_price_values(row: list[str], header: list[str]) -> list[float]:
    normalized_header = [_clean_text(cell).lower() for cell in header]
    price_indices = [
        index
        for index, cell_name in enumerate(normalized_header)
        if any(token in cell_name for token in ("price", "cost", "цена", "стоим", "тариф", "ндс", "sum"))
    ]
    values: list[float] = []
    if price_indices:
        for index in price_indices:
            candidate = _clean_text(row[index] if index < len(row) else None)
            price = _parse_price_value_v2(candidate)
            if price is not None:
                values.append(price)
        if values:
            return values

    start_index = 0
    if header:
        name_indices = _header_name_indices(header)
        if name_indices:
            start_index = max(name_indices) + 1
        elif len(row) >= 4:
            start_index = 4
    elif len(row) >= 4:
        start_index = 3

    numeric_cells = [
        value
        for value in row[start_index:]
        if re.search(r"\d", _clean_text(value))
    ]
    if not numeric_cells:
        return []

    for value in numeric_cells:
        price = _parse_price_value_v2(value)
        if price is not None:
            values.append(price)
    return values


def _extract_service_name_from_row(row: list[str], code: str | None, prices: list[float], header: list[str]) -> str | None:
    cleaned = [_clean_text(cell) for cell in row if _clean_text(cell)]
    if not cleaned:
        return None
    row_text = " ".join(cleaned)
    if is_section_row(row_text) and not prices:
        return None

    # Structured spreadsheets usually keep the service title in one of the
    # first text cells after the code column. Keep this permissive so older
    # XLS files do not lose valid service names.
    direct_candidates: list[str] = []
    for cell in row[:6]:
        candidate = _clean_text(cell)
        if not candidate or is_price_like(candidate):
            continue
        if re.fullmatch(r"\d+(?:\.\d+)?", candidate):
            continue
        if code and _normalize_code_token(candidate) == code:
            continue
        candidate = _strip_leading_row_noise(candidate)
        if not candidate or is_section_row(candidate):
            continue
        alnum_count = sum(1 for ch in candidate if ch.isalnum())
        if alnum_count < 3:
            continue
        direct_candidates.append(candidate)
    if direct_candidates:
        direct_candidates.sort(key=len, reverse=True)
        return direct_candidates[0]

    pieces: list[str] = []
    for cell in cleaned:
        normalized_cell = _clean_text(cell)
        if code and _normalize_code_token(normalized_cell) == code:
            continue
        if is_price_like(normalized_cell):
            continue
        text = _strip_leading_row_noise(normalized_cell)
        if code and text.startswith(code):
            text = _clean_text(text[len(code):])
        else:
            tokens = text.split()
            if tokens:
                first = _normalize_code_token(tokens[0])
                if first == code or is_code_like(first):
                    text = _clean_text(" ".join(tokens[1:]))
        if text and not is_price_like(text):
            pieces.append(text)

    name = _clean_text(" ".join(pieces))
    if not name:
        return None
    if code and name == code:
        return None
    if len(name) < 3:
        return None
    if re.fullmatch(r"[\d\s.,-]+", name):
        return None
    return name


def _infer_partner_name_from_path(file_path: Path) -> str:
    stem = file_path.stem
    match = re.match(r"^(Клиника\s*\d+)", stem, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()
    normalized = _clean_partner_name_from_stem(stem)
    if normalized != stem:
        return normalized
    match = re.match(r"^(Клиника\s+\d+)", stem, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()

    cleaned = re.sub(r"(прайс|price|год)", " ", stem, flags=re.IGNORECASE)
    cleaned = re.sub(r"\b\d{4}\b", " ", cleaned)
    cleaned = re.sub(r"[_-]+", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ._-")
    return cleaned or stem

def _entry_partner_name(archive_name: str, entry_path: PurePosixPath) -> str:
    file_partner_name = _infer_partner_name_from_path(Path(entry_path.name))
    if file_partner_name and file_partner_name != entry_path.stem:
        return file_partner_name
    group_key = _group_key(entry_path)
    return _choose_partner_name(archive_name, group_key)


def _current_price_value(
    price_resident: float | None,
    price_nonresident: float | None,
    price_original: float | None,
) -> float | None:
    for value in (price_resident, price_original, price_nonresident):
        if value is not None:
            return float(value)
    return None


def _previous_price_item(
    session: Session,
    *,
    partner_id,
    service_id: str | None,
    document_date: date | None,
) -> PriceItem | None:
    if service_id is None:
        return None
    query = select(PriceItem).where(
        PriceItem.partner_id == partner_id,
        PriceItem.service_id == service_id,
        PriceItem.is_active.is_(True),
    )
    if document_date is not None:
        query = query.where(
            (PriceItem.effective_date.is_(None)) | (PriceItem.effective_date < document_date)
        )
    return session.scalars(
        query.order_by(PriceItem.effective_date.desc().nullslast())
    ).first()


def _price_change_note(previous_item: PriceItem | None, current_price: float | None) -> str | None:
    if previous_item is None or current_price is None:
        return None
    previous_price = _current_price_value(
        float(previous_item.price_resident_kzt) if previous_item.price_resident_kzt is not None else None,
        float(previous_item.price_nonresident_kzt) if previous_item.price_nonresident_kzt is not None else None,
        float(previous_item.price_original) if previous_item.price_original is not None else None,
    )
    if previous_price is None or previous_price <= 0:
        return None
    change = abs(current_price - previous_price) / previous_price
    if change <= REVIEW_PRICE_CHANGE_THRESHOLD:
        return None
    percent = round(change * 100)
    return f"anomaly: price changed by {percent}% from previous version"


def _archive_superseded_items(
    session: Session,
    *,
    partner_id,
    service_id: str | None,
    document_date: date | None,
) -> None:
    if service_id is None or document_date is None:
        return
    stmt = select(PriceItem).where(
        PriceItem.partner_id == partner_id,
        PriceItem.service_id == service_id,
        PriceItem.is_active.is_(True),
    )
    if document_date is not None:
        stmt = stmt.where(
            (PriceItem.effective_date.is_(None)) | (PriceItem.effective_date <= document_date)
        )
    duplicates = session.scalars(stmt).all()
    for duplicate in duplicates:
        duplicate.is_active = False


def _store_document(
    session: Session,
    *,
    archive: ArchiveBatch | None,
    partner: Partner,
    file_path: Path,
    data: bytes,
    stats: ArchiveImportStats,
    service_choices: list[tuple[str, str]],
    source_path: str | None = None,
    display_name: str | None = None,
) -> PriceDocument:
    suffix = file_path.suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        stats.warnings.append(f"Skipped unsupported file: {file_path.name}")
        return

    document_date = _extract_date(file_path.stem) or _extract_date(file_path.parent.name) or None
    try:
        rows, content = extract_rows_and_content(file_path, data, stats)
    except Exception as exc:
        stats.warnings.append(f"Failed to read {file_path.name}: {exc}")
        rows, content = [], ""

    parsed_rows = _rows_from_source(rows)
    has_readable_content = bool(_clean_text(content))
    parse_status = "processed" if parsed_rows else "error"
    if not parsed_rows and has_readable_content:
        parse_status = "needs_review"
    needs_review = parse_status == "needs_review"
    if document_date and document_date > date.today():
        stats.warnings.append(f"Price date is in the future: {file_path.name} -> {document_date.isoformat()}")
        needs_review = True

    document = PriceDocument(
        archive_id=archive.archive_id if archive else None,
        partner_id=partner.partner_id,
        file_name=display_name or file_path.name,
        source_path=source_path or display_name or file_path.name,
        file_format=suffix.lstrip("."),
        effective_date=document_date,
        parsed_at=datetime.utcnow() if parsed_rows or content else None,
        parse_status=parse_status,
        parse_log="\n".join(stats.warnings[-5:]) if stats.warnings else None,
        raw_content=content[:50000] if content else None,
    )
    session.add(document)
    session.flush()
    stats.document_count += 1

    for row in parsed_rows:
        raw_name = _clean_text(row.get("service_name_raw"))
        if not raw_name:
            stats.warnings.append(f"Skipped row without service name: {file_path.name}")
            continue
        service_id, score, match_status = _best_service_match(str(raw_name), service_choices)
        if service_id:
            stats.matched_item_count += 1
        row_text = _clean_text(row.get("raw_row"))
        currency_original = _detect_currency(row_text)
        raw_price_resident = row["price_resident_kzt"] if isinstance(row.get("price_resident_kzt"), float) else None
        raw_price_nonresident = row["price_nonresident_kzt"] if isinstance(row.get("price_nonresident_kzt"), float) else None
        raw_price_original = row["price_original"] if isinstance(row.get("price_original"), float) else raw_price_resident
        price_resident, resident_rate_source = _convert_price_to_kzt(raw_price_resident, currency_original, stats)
        price_nonresident, nonresident_rate_source = _convert_price_to_kzt(raw_price_nonresident, currency_original, stats)
        price_original = raw_price_original
        rate_source = resident_rate_source if raw_price_resident is not None else nonresident_rate_source
        previous_item = _previous_price_item(
            session,
            partner_id=partner.partner_id,
            service_id=service_id,
            document_date=document_date,
        )
        anomaly_note = _price_change_note(
            previous_item,
            _current_price_value(price_resident, price_nonresident, price_original),
        )
        _archive_superseded_items(
            session,
            partner_id=partner.partner_id,
            service_id=service_id,
            document_date=document_date,
        )
        if (
            match_status == "fuzzy_review"
            or _row_needs_review(
            service_id=service_id,
            price_resident=price_resident,
            price_nonresident=price_nonresident,
            price_original=price_original,
        )
            or anomaly_note
        ):
            needs_review = True
        validation_reasons = _price_validation_reasons(
            price_resident=price_resident,
            price_nonresident=price_nonresident,
            price_original=price_original,
        )
        note_parts: list[str] = []
        if service_id is None:
            if match_status == "fuzzy_review":
                note_parts.append(f"service review suggested ({score}%)")
            else:
                note_parts.append("service match required")
        elif match_status == "exact":
            note_parts.append("exact matched")
        elif match_status == "synonym":
            note_parts.append("synonym matched")
        elif match_status == "fuzzy_auto":
            note_parts.append(f"auto matched ({score}%)")
        note_parts.extend(validation_reasons)
        if currency_original != "KZT":
            note_parts.append(f"currency converted from {currency_original} via {rate_source} rate")
        if document_date and document_date > date.today():
            note_parts.append("price date is in the future")
        if anomaly_note:
            note_parts.append(anomaly_note)
        verification_note = "; ".join(dict.fromkeys(part for part in note_parts if part))
        item = PriceItem(
            doc_id=document.doc_id,
            partner_id=partner.partner_id,
            service_name_raw=_truncate_text(raw_name, 500) or file_path.stem,
            service_code_source=_truncate_text(row["service_code_source"], 128),
            service_id=service_id,
            price_resident_kzt=price_resident,
            price_nonresident_kzt=price_nonresident,
            price_original=price_original,
            currency_original=currency_original if any(value is not None for value in (price_resident, price_nonresident, price_original)) else None,
            is_verified=False,
            verification_note=_truncate_text(verification_note, 1000),
            effective_date=document_date,
            is_active=True,
        )
        session.add(item)
        stats.item_count += 1

    if not parsed_rows and document.parse_status != "error":
        item = PriceItem(
            doc_id=document.doc_id,
            partner_id=partner.partner_id,
            service_name_raw=_truncate_text(file_path.stem, 500) or file_path.name,
            service_code_source=None,
            service_id=None,
            price_resident_kzt=None,
            price_nonresident_kzt=None,
            price_original=None,
            currency_original=None,
            is_verified=False,
            verification_note="needs manual review",
            effective_date=document_date,
            is_active=True,
        )
        session.add(item)
        stats.item_count += 1

    if document.parse_status != "error":
        document.parse_status = "needs_review" if needs_review else "processed"
    return document


def import_single_document(
    session: Session,
    file_path: Path,
    data: bytes,
    original_name: str | None = None,
) -> tuple[ArchiveImportStats, PriceDocument]:
    stats = ArchiveImportStats(archive_name=file_path.stem)
    service_choices = _build_service_choices(session)
    partner_name = _infer_partner_name_from_path(file_path)
    partner = session.scalars(select(Partner).where(Partner.name == partner_name)).first()
    if partner is None:
        partner = Partner(name=partner_name, is_active=True)
        session.add(partner)
        session.flush()
        stats.partner_count += 1
    document = _store_document(
        session,
        archive=None,
        partner=partner,
        file_path=file_path,
        data=data,
        stats=stats,
        service_choices=service_choices,
        source_path=original_name,
        display_name=original_name,
    )
    session.commit()
    return stats, document


def import_archive(session: Session, archive_path: Path) -> ArchiveImportStats:
    archive_name = archive_path.stem
    archive_record = ArchiveBatch(
        file_name=archive_path.name,
        saved_path=str(archive_path),
        status="processing",
        warnings=[],
    )
    session.add(archive_record)
    session.flush()

    stats = ArchiveImportStats(archive_name=archive_name, archive_id=str(archive_record.archive_id))
    existing_partners = {
        partner.name.strip().lower(): partner
        for partner in session.scalars(select(Partner)).all()
        if partner.name.strip()
    }
    service_choices = _build_service_choices(session)

    with zipfile.ZipFile(archive_path) as zf:
        entries = [
            info
            for info in zf.infolist()
            if not info.is_dir()
            and info.filename
            and not any(part.lower() in IGNORED_PATH_PARTS for part in PurePosixPath(info.filename).parts)
        ]

        grouped_entries: dict[str, list[zipfile.ZipInfo]] = {}
        for info in entries:
            normalized_path = _normalize_path(info.filename)
            if not normalized_path.parts:
                continue
            grouped_entries.setdefault(_group_key(normalized_path), []).append(info)

        for group_infos in grouped_entries.values():
            for info in sorted(group_infos, key=lambda item: item.filename):
                normalized_path = _normalize_path(info.filename)
                partner_name = _entry_partner_name(archive_name, normalized_path)
                partner = existing_partners.get(partner_name.lower())
                if partner is None:
                    partner = Partner(name=partner_name, is_active=True)
                    session.add(partner)
                    session.flush()
                    existing_partners[partner_name.lower()] = partner
                    stats.partner_count += 1
                data = zf.read(info)
                _store_document(
                    session,
                    archive=archive_record,
                    partner=partner,
                    file_path=Path(normalized_path.name),
                    data=data,
                    stats=stats,
                    service_choices=service_choices,
                    source_path=str(normalized_path),
                )

    archive_record.partner_count = stats.partner_count
    archive_record.document_count = stats.document_count
    archive_record.item_count = stats.item_count
    archive_record.matched_item_count = stats.matched_item_count
    archive_documents = list(
        session.scalars(
            select(PriceDocument).where(PriceDocument.archive_id == archive_record.archive_id)
        ).all()
    )
    if any(document.parse_status == "error" for document in archive_documents):
        archive_record.status = "error"
    elif not stats.document_count or any(document.parse_status == "needs_review" for document in archive_documents):
        archive_record.status = "needs_review"
    else:
        archive_record.status = "processed"
    archive_record.processed_at = datetime.utcnow()
    archive_record.warnings = stats.warnings
    session.commit()
    return stats


def import_directory(session: Session, directory: Path) -> ArchiveImportStats:
    stats = ArchiveImportStats(archive_name=directory.name)
    service_choices = _build_service_choices(session)
    existing_partners = {
        partner.name.strip().lower(): partner
        for partner in session.scalars(select(Partner)).all()
        if partner.name.strip()
    }

    for file_path in sorted(directory.iterdir()):
        if not file_path.is_file() or file_path.suffix.lower() == ".zip":
            continue
        partner_name = _infer_partner_name_from_path(file_path)
        partner = existing_partners.get(partner_name.lower())
        if partner is None:
            partner = Partner(name=partner_name, is_active=True)
            session.add(partner)
            session.flush()
            existing_partners[partner_name.lower()] = partner
            stats.partner_count += 1

        _store_document(
            session,
            archive=None,
            partner=partner,
            file_path=file_path,
            data=file_path.read_bytes(),
            stats=stats,
            service_choices=service_choices,
        )

    session.commit()
    return stats

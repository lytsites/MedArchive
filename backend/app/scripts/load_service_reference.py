from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.core.database import SessionLocal, init_db
from app.models import Service


def _header_index(headers: list[str], *names: str) -> int:
    for name in names:
        if name in headers:
            return headers.index(name)
    raise ValueError(f"Missing expected header. Available headers: {headers}")


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python -m app.scripts.load_service_reference <path-to-xlsx>")

    path = Path(sys.argv[1])
    if not path.exists():
        raise SystemExit(f"File not found: {path}")

    init_db()
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("Imported 0 services from reference.")
        return

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = rows[1:]

    specialty_header = "\u0421\u043f\u0435\u0446\u0438\u0430\u043b\u044c\u043d\u043e\u0441\u0442\u044c"
    name_idx = _header_index(headers, "Name_ru")
    specialty_idx = _header_index(headers, specialty_header, "Specialty", "specialty")
    tariff_idx = _header_index(headers, "TarificatrCode", "TariffCode", "tariff_code")

    session = SessionLocal()
    try:
        existing_names = set(session.scalars(select(Service.service_name)).all())
        created = 0
        for row in data_rows:
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            service_name = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] is not None else ""
            if not service_name or service_name in existing_names:
                continue

            specialty = row[specialty_idx] if specialty_idx < len(row) else None
            tariff = row[tariff_idx] if tariff_idx < len(row) else None

            synonyms = []
            if specialty is not None and str(specialty).strip():
                synonyms.append(str(specialty).strip())

            service = Service(
                service_name=service_name,
                category=str(specialty).strip() if specialty is not None else None,
                icd_code=str(tariff).strip() if tariff is not None else None,
                synonyms=synonyms or None,
            )
            session.add(service)
            existing_names.add(service_name)
            created += 1

        session.commit()
        print(f"Imported {created} services from reference.")
    finally:
        session.close()


if __name__ == "__main__":
    main()
